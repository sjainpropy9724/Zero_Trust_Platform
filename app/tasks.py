from celery import Celery
from app.db.database import SessionLocal
from app.db import models
from app.services import nlp_service
from pathlib import Path
import docx
import openpyxl
from pypdf import PdfReader

celery_app = Celery("tasks", broker="redis://redis:6379/0")

def get_text_from_docx(filepath: Path) -> str:
    """Extracts text from a .docx file."""
    try:
        doc = docx.Document(filepath)
        return "\n".join([para.text for para in doc.paragraphs])
    except Exception as e:
        print(f"Error reading docx file {filepath}: {e}")
        return ""

def get_text_from_pdf(filepath: Path) -> str:
    """Extracts text from a .pdf file."""
    try:
        reader = PdfReader(filepath)
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"
        return text
    except Exception as e:
        print(f"Error reading pdf file {filepath}: {e}")
        return ""

def get_text_from_xlsx(filepath: Path) -> str:
    """Extracts text from an .xlsx file."""
    try:
        workbook = openpyxl.load_workbook(filepath)
        text = ""
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        text += str(cell.value) + " "
                text += "\n"
        return text
    except Exception as e:
        print(f"Error reading xlsx file {filepath}: {e}")
        return ""

@celery_app.task
def scan_file_task(file_id: int):
    """
    Celery task to scan a file for PII in the background.
    """
    print(f"Starting PII scan for file_id: {file_id}")
    db = SessionLocal()
    try:
        db_file = db.query(models.File).filter(models.File.id == file_id).first()
        if not db_file:
            print(f"File with id {file_id} not found.")
            return

        file_path = Path(db_file.filepath)
        if not file_path.exists():
            print(f"File not found on disk: {file_path}")
            return
        
        found_pii_overall = []

        file_suffix = file_path.suffix.lower()

        if file_suffix in [".txt", ".docx", ".pdf"]:
            content_to_scan = ""
            if file_suffix == ".txt":
                content_to_scan = file_path.read_text(encoding='utf-8', errors='ignore')
            elif file_suffix == ".docx":
                content_to_scan = get_text_from_docx(file_path)
            elif file_suffix == ".pdf":
                content_to_scan = get_text_from_pdf(file_path)
            
            if content_to_scan.strip():
                found_pii_overall = nlp_service.scan_text_for_pii(content_to_scan)

        elif file_suffix == ".xlsx":
            workbook = openpyxl.load_workbook(file_path)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value:
                            pii_in_cell = nlp_service.scan_text_for_pii(str(cell.value))
                            for pii in pii_in_cell:
                                pii["location"] = f"Sheet '{sheet_name}', Cell {cell.coordinate}"
                                found_pii_overall.append(pii)
        else:
            print(f"Skipping scan for unsupported file type: {file_path.suffix}")
            db_file.sensitivity_level = "Unsupported"
            db.commit()
            return

        if not found_pii_overall:
            db_file.sensitivity_level = "Clean"
        else:
            high_risk_types = {
                'PERSON', 'EMAIL', 'PHONE_NUMBER', 'CREDIT_CARD', 
                'SSN', 'US_DRIVER_LICENSE', 'MEDICAL_LICENSE', 'IBAN_CODE'
            }
            medium_risk_types = {'LOCATION', 'ORGANIZATION', 'IP_ADDRESS', 'URL', 'DATE_TIME'}

            is_confidential = any(entity['type'] in high_risk_types for entity in found_pii_overall)
            is_internal = any(entity['type'] in medium_risk_types for entity in found_pii_overall)
            if is_confidential:
                db_file.sensitivity_level = "Confidential"
            elif is_internal:
                db_file.sensitivity_level = "Internal"
            else:
                db_file.sensitivity_level = "Internal"

        print(f"Scan complete for file {file_id}. Level: {db_file.sensitivity_level}. PII: {found_pii_overall}")
        db.commit()

    finally:
        db.close()